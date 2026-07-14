"""
JPDC CRM Backend - Database Setup & Import
Jinxi Seaview City (金禧海景城) Sales Management System
"""
import openpyxl, sqlite3, os, sys
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jpdc.db")

def clean(v):
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = str(v).strip()
    if s in ("", "—", "N/A", "None", "nan"): return None
    try: return float(s)
    except: pass
    return s

def create_tables(conn):
    cur = conn.cursor()
    
    # Users & Roles
    cur.execute("""CREATE TABLE IF NOT EXISTS roles (
        id INTEGER PRIMARY KEY, name TEXT UNIQUE NOT NULL,
        can_view_all INTEGER DEFAULT 1, can_edit INTEGER DEFAULT 0,
        can_manage_users INTEGER DEFAULT 0, can_export INTEGER DEFAULT 0,
        can_view_financial INTEGER DEFAULT 0
    )""")
    
    cur.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY, username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL, role_id INTEGER NOT NULL DEFAULT 3,
        display_name TEXT, active INTEGER DEFAULT 1,
        created_at TEXT, last_login TEXT,
        FOREIGN KEY(role_id) REFERENCES roles(id)
    )""")
    
    # Default roles
    cur.execute("SELECT COUNT(*) FROM roles")
    if cur.fetchone()[0] == 0:
        roles = [
            ("Admin", 1, 1, 1, 1, 1),
            ("Manager", 1, 1, 0, 1, 1),
            ("Agent", 1, 0, 0, 0, 0),
            ("Viewer", 1, 0, 0, 0, 0),
        ]
        cur.executemany("INSERT INTO roles VALUES (?,?,?,?,?,?,?)",
            [(i+1, r[0], r[1], r[2], r[3], r[4], r[5]) for i, r in enumerate(roles)])
        cur.execute("INSERT INTO users VALUES (1,'admin','admin123',1,'Administrator',1,?,NULL)", 
            (datetime.now().isoformat(),))

    # Units (combined Tower A + E)
    cur.execute("""CREATE TABLE IF NOT EXISTS units (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tower TEXT NOT NULL, floor_num TEXT, unit_no TEXT NOT NULL,
        unit_type TEXT, area_sqm REAL, old_area_sqm REAL,
        price_per_sqm REAL, total_price REAL,
        discount REAL, discounted_price REAL,
        status TEXT, situation TEXT,
        buyer_remarks TEXT, sales_date TEXT,
        remarks TEXT, sd TEXT, sm TEXT, pia TEXT,
        created_at TEXT, updated_at TEXT
    )""")
    
    # Overdue warnings
    cur.execute("""CREATE TABLE IF NOT EXISTS overdue_warnings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        buyer_name TEXT, tower TEXT, unit TEXT,
        paid_installments INTEGER, total_paid REAL,
        days_stopped INTEGER, risk_level TEXT,
        notes TEXT, created_at TEXT
    )""")
    
    # Sold clients detail
    cur.execute("""CREATE TABLE IF NOT EXISTS sold_clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tower TEXT, floor_num TEXT, unit_no TEXT,
        unit_type TEXT, area_sqm REAL, total_price REAL,
        buyer_name TEXT, sd TEXT, sm TEXT, pia TEXT,
        sales_date TEXT, payment_status TEXT,
        notes TEXT, created_at TEXT
    )""")
    
    # Payment details
    cur.execute("""CREATE TABLE IF NOT EXISTS payment_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        buyer_name TEXT, tower TEXT, unit_no TEXT,
        unit_type TEXT, sales_date TEXT,
        tr1_date TEXT, tr1_amount REAL,
        tr2_date TEXT, tr2_amount REAL,
        tr3_date TEXT, tr3_amount REAL,
        tr4_date TEXT, tr4_amount REAL,
        tr5_date TEXT, tr5_amount REAL,
        tr6_date TEXT, tr6_amount REAL,
        tr7_date TEXT, tr7_amount REAL,
        tr8_date TEXT, tr8_amount REAL,
        total_paid REAL, created_at TEXT
    )""")
    
    # Returned units
    cur.execute("""CREATE TABLE IF NOT EXISTS returned_units (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tower TEXT, floor_num TEXT, unit_no TEXT,
        unit_type TEXT, area_sqm REAL, price REAL,
        buyer_notes TEXT, sd TEXT, sm TEXT, pia TEXT,
        return_reason TEXT, created_at TEXT
    )""")
    
    # Problem/reserved units
    cur.execute("""CREATE TABLE IF NOT EXISTS problem_units (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tower TEXT, floor_num TEXT, unit_no TEXT,
        unit_type TEXT, area_sqm REAL, price REAL,
        buyer_notes TEXT, sd TEXT, sm TEXT, pia TEXT,
        issue_status TEXT, created_at TEXT
    )""")
    
    # Unit type heatmap / inventory summary
    cur.execute("""CREATE TABLE IF NOT EXISTS unit_inventory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tower TEXT, unit_type TEXT,
        available INTEGER DEFAULT 0, sold INTEGER DEFAULT 0,
        returned INTEGER DEFAULT 0, reserved INTEGER DEFAULT 0,
        on_hold INTEGER DEFAULT 0, late INTEGER DEFAULT 0,
        total INTEGER DEFAULT 0, updated_at TEXT
    )""")
    
    # Sync meta
    cur.execute("""CREATE TABLE IF NOT EXISTS sync_meta (
        key TEXT PRIMARY KEY, value TEXT, updated_at TEXT
    )""")
    
    conn.commit()

# ──── Import Functions ──────────────────────────────────────

def get_sheet(wb, name):
    for s in wb.sheetnames:
        if name in str(s):
            return wb[s]
    return None

def import_units(conn, wb, tower_name):
    ws = get_sheet(wb, tower_name)
    if not ws: return 0
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for i, row in enumerate(rows[1:], 2):
        vals = [clean(row[j]) if j < len(row) else None for j in range(14)]
        if vals[1] is None: continue
        data.append({
            'tower': tower_name,
            'floor_num': str(vals[0]) if vals[0] else '',
            'unit_no': str(vals[1]) if vals[1] else '',
            'unit_type': str(vals[2]) if vals[2] else '',
            'area_sqm': vals[3],
            'old_area_sqm': vals[4],
            'price_per_sqm': vals[5],
            'total_price': vals[6],
            'discount': vals[7],
            'discounted_price': vals[8],
            'status': str(vals[9]) if vals[9] else '',
            'situation': str(vals[10]) if len(vals) > 10 and vals[10] else None,
            'buyer_remarks': str(vals[11]) if len(vals) > 11 and vals[11] else None,
            'sales_date': str(vals[12]) if len(vals) > 12 and vals[12] else None,
            'remarks': str(vals[13]) if len(vals) > 13 and vals[13] else None,
        })
    
    cur = conn.cursor()
    cur.execute("DELETE FROM units WHERE tower = ?", (tower_name,))
    for d in data:
        cur.execute("""INSERT INTO units (tower,floor_num,unit_no,unit_type,area_sqm,old_area_sqm,
            price_per_sqm,total_price,discount,discounted_price,status,situation,buyer_remarks,sales_date,remarks)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d['tower'],d['floor_num'],d['unit_no'],d['unit_type'],d['area_sqm'],d['old_area_sqm'],
             d['price_per_sqm'],d['total_price'],d['discount'],d['discounted_price'],
             d['status'],d['situation'],d['buyer_remarks'],d['sales_date'],d['remarks']))
    conn.commit()
    return len(data)

def import_overdue(conn, wb):
    ws = get_sheet(wb, "逾期预警")
    if not ws: return 0
    rows = list(ws.iter_rows(values_only=True))
    cur = conn.cursor()
    cur.execute("DELETE FROM overdue_warnings")
    count = 0
    for row in rows[1:]:
        if not row or not row[1]: continue
        cur.execute("""INSERT INTO overdue_warnings (buyer_name,tower,unit,paid_installments,total_paid,days_stopped,risk_level)
            VALUES (?,?,?,?,?,?,?)""",
            (str(row[1]),str(row[2]),str(row[3]),int(row[4] or 0),float(str(row[5]).replace(',','')) if row[5] else 0,
             int(row[6] or 0),str(row[7]) if len(row)>7 and row[7] else ''))
        count += 1
    conn.commit()
    return count

def import_sold_clients(conn, wb):
    ws = get_sheet(wb, "已售客户")
    if not ws: return 0
    rows = list(ws.iter_rows(values_only=True))
    cur = conn.cursor()
    cur.execute("DELETE FROM sold_clients")
    count = 0
    for row in rows[3:]:  # Skip 3 header rows
        if not row or len(row) < 7 or not row[2]: continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(12)]
        if not vals[1]: continue
        cur.execute("""INSERT INTO sold_clients (tower,floor_num,unit_no,unit_type,area_sqm,total_price,buyer_name,sd,sm,pia)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (str(vals[0]) if vals[0] else '', str(vals[1]) if vals[1] else '',
             str(vals[2]) if vals[2] else '', str(vals[3]) if vals[3] else '',
             vals[4], vals[5],
             str(vals[6])[:200] if vals[6] else '',
             str(vals[7]) if vals[7] else None,
             str(vals[8]) if vals[8] else None,
             str(vals[9]) if vals[9] else None))
        count += 1
    conn.commit()
    return count

def import_payments(conn, wb):
    ws = get_sheet(wb, "Payment")
    if not ws: return 0
    rows = list(ws.iter_rows(values_only=True))
    cur = conn.cursor()
    cur.execute("DELETE FROM payment_details")
    count = 0
    for row in rows[1:]:
        if not row or not row[0] or str(row[0]).startswith("合计"): continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(32)]
        cur.execute("""INSERT INTO payment_details (buyer_name,tower,unit_no,unit_type,sales_date,
            tr1_date,tr1_amount,tr2_date,tr2_amount,tr3_date,tr3_amount,tr4_date,tr4_amount,
            tr5_date,tr5_amount,tr6_date,tr6_amount,tr7_date,tr7_amount,tr8_date,tr8_amount)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (str(vals[0])[:100], str(vals[1])[:5], str(vals[3]) if vals[3] else '', '',
             str(vals[4]) if vals[4] else '',
             str(vals[5]) if vals[5] else '', vals[6],
             str(vals[7]) if vals[7] else '', vals[8],
             str(vals[9]) if vals[9] else '', vals[10],
             str(vals[11]) if vals[11] else '', vals[12],
             str(vals[13]) if vals[13] else '', vals[14],
             str(vals[15]) if vals[15] else '', vals[16],
             str(vals[17]) if vals[17] else '', vals[18],
             str(vals[19]) if vals[19] else '', vals[20]))
        count += 1
    conn.commit()
    return count

def import_returned(conn, wb):
    ws = get_sheet(wb, "退房明细")
    if not ws: return 0
    rows = list(ws.iter_rows(values_only=True))
    cur = conn.cursor()
    cur.execute("DELETE FROM returned_units")
    count = 0
    for row in rows[4:]:
        if not row or not row[2]: continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(11)]
        cur.execute("""INSERT INTO returned_units (tower,floor_num,unit_no,unit_type,area_sqm,price,buyer_notes,sd,sm,pia)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (str(vals[0]) if vals[0] else '', str(vals[1]) if vals[1] else '',
             str(vals[2]) if vals[2] else '', str(vals[3]) if vals[3] else '',
             vals[4], vals[5],
             str(vals[6])[:200] if vals[6] else '',
             str(vals[7]) if vals[7] else None,
             str(vals[8]) if vals[8] else None,
             str(vals[9]) if vals[9] else None))
        count += 1
    conn.commit()
    return count

def import_problem(conn, wb):
    ws = get_sheet(wb, "问题单元")
    if not ws: return 0
    rows = list(ws.iter_rows(values_only=True))
    cur = conn.cursor()
    cur.execute("DELETE FROM problem_units")
    count = 0
    for row in rows[4:]:
        if not row or not row[2]: continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(12)]
        cur.execute("""INSERT INTO problem_units (tower,floor_num,unit_no,unit_type,area_sqm,price,buyer_notes,sd,sm,pia,issue_status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (str(vals[0]) if vals[0] else '', str(vals[1]) if vals[1] else '',
             str(vals[2]) if vals[2] else '', str(vals[3]) if vals[3] else '',
             vals[4], vals[5],
             str(vals[6])[:200] if vals[6] else '',
             str(vals[7]) if vals[7] else None,
             str(vals[8]) if vals[8] else None,
             str(vals[9]) if vals[9] else None,
             str(vals[10]) if vals[10] else None))
        count += 1
    conn.commit()
    return count

def import_inventory(conn, wb):
    """Build inventory summary from units table"""
    cur = conn.cursor()
    cur.execute("DELETE FROM unit_inventory")
    
    for tower in ['A', 'E']:
        for utype in ['STUDIO', '2 BEDROOM', '3 BEDROOM']:
            cur.execute("SELECT COUNT(*) FROM units WHERE tower=? AND unit_type=?", (tower, utype))
            total = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM units WHERE tower=? AND unit_type=? AND status='可售'", (tower, utype))
            avail = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM units WHERE tower=? AND unit_type=? AND status IN ('已售','Sold')", (tower, utype))
            sold = cur.fetchone()[0]
            
            cur.execute("""INSERT INTO unit_inventory (tower,unit_type,available,sold,returned,reserved,on_hold,late,total,updated_at)
                VALUES (?,?,?,?,0,0,0,0,?,?)""", (tower, utype, avail, sold, total, datetime.now().isoformat()))
    
    # E tower also has 4 BEDROOM
    for utype in ['4 BEDROOM']:
        cur.execute("SELECT COUNT(*) FROM units WHERE tower='E' AND unit_type=?", (utype,))
        total = cur.fetchone()[0]
        if total > 0:
            cur.execute("SELECT COUNT(*) FROM units WHERE tower='E' AND unit_type=? AND status='可售'", (utype,))
            avail = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM units WHERE tower='E' AND unit_type=? AND status IN ('已售','Sold')", (utype,))
            sold = cur.fetchone()[0]
            cur.execute("""INSERT INTO unit_inventory (tower,unit_type,available,sold,returned,reserved,on_hold,late,total,updated_at)
                VALUES (?,?,?,?,0,0,0,0,?,?)""", ('E', utype, avail, sold, total, datetime.now().isoformat()))
    conn.commit()
    return 1

# ──── Main ──────────────────────────────────────────────────

def main(excel_path=None):
    if not excel_path:
        excel_path = r"D:\2026年7月\CRM\JPDC_CRM-7.6_linked_v2_(5)_backup_20260706_1506.xlsx"
    
    print(f"{'='*55}")
    print(f"  JPDC CRM Database Builder")
    print(f"  Excel: {os.path.basename(excel_path)}")
    print(f"  DB: {DB_PATH}")
    print(f"{'='*55}\n")
    
    conn = sqlite3.connect(DB_PATH)
    create_tables(conn)
    
    print("Loading Excel...")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    
    results = {}
    
    print("Importing units...")
    results['Tower A Units'] = import_units(conn, wb, 'TOWER A')
    results['Tower E Units'] = import_units(conn, wb, 'TOWER E')
    
    print("Importing overdue warnings...")
    results['Overdue'] = import_overdue(conn, wb)
    
    print("Importing sold clients...")
    results['Sold'] = import_sold_clients(conn, wb)
    
    print("Importing payments...")
    results['Payments'] = import_payments(conn, wb)
    
    print("Importing returned units...")
    results['Returned'] = import_returned(conn, wb)
    
    print("Importing problem units...")
    results['Problem'] = import_problem(conn, wb)
    
    print("Building inventory...")
    results['Inventory'] = import_inventory(conn, wb)
    
    # Store sync meta
    cur = conn.cursor()
    cur.execute("INSERT OR REPLACE INTO sync_meta (key, value, updated_at) VALUES (?,?,?)",
        ('last_import', datetime.now().isoformat(), datetime.now().isoformat()))
    conn.commit()
    
    wb.close()
    conn.close()
    
    total = sum(r for r in results.values() if isinstance(r, int))
    print(f"\n{'='*55}")
    for k, v in results.items():
        print(f"  {k:<20}: {v}")
    print(f"  {'─'*35}")
    print(f"  {'TOTAL':<20}: {total}")
    print(f"{'='*55}")
    print(f"\n  DB ready at: {DB_PATH}")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    main(path)
