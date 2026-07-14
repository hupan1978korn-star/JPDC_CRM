"""
JPDC CRM - Excel Import / Monthly Update Script
Auto-detect header rows, backup, validate, log
Usage: python import_excel.py [path/to/excel.xlsx]
       Double-click 一键更新数据.bat
"""
import openpyxl, sqlite3, os, sys, glob, shutil, json
from datetime import datetime

# ── Config ──────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jpdc.db")
LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "import_log.txt")
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_data")
SOURCE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "source_excels")
os.makedirs(SOURCE_DIR, exist_ok=True)

EXCEL_SEARCH_PATHS = [
    os.path.join(os.path.expanduser("~/Desktop"), "JPDC_CRM*.xlsx"),
    os.path.join(os.path.expanduser("~/Desktop"), "*.xlsx"),
    r"D:\2026年7月\CRM\*.xlsx",
]

# Header keywords for auto-detection (row offsets)
SHEET_PATTERNS = {
    "TOWER A":   {"sheets": ["TOWER A"], "table": "units", "tower_val": "A", "clear_col": "tower", "clear_vals": ["A","TOWER A"]},
    "TOWER E":   {"sheets": ["TOWER E"], "table": "units", "tower_val": "E", "clear_col": "tower", "clear_vals": ["E","TOWER E"]},
    "Overdue":   {"sheets": ["Overdue"], "table": "overdue_warnings", "clear": True},
    "Sold":      {"sheets": ["Sold"], "table": "sold_clients", "clear": True},
    "Payments":  {"sheets": ["Payment"], "table": "payment_details", "clear": True},
    "Returned":  {"sheets": ["Returned"], "table": "returned_units", "clear": True},
    "Problem":   {"sheets": ["Problem"], "table": "problem_units", "clear": True},
}

# ── Helpers ─────────────────────────────────────────────────
def log(msg, level="INFO"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [{level}] {msg}"
    print(line)
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except:
        pass

def clean(v):
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = str(v).strip()
    if s in ("", "—", "N/A", "None", "nan"): return None
    try: return float(s)
    except: pass
    return s

def find_excel():
    log("Searching for Excel files...")
    seen = set()
    candidates = []
    for pat in EXCEL_SEARCH_PATHS:
        for f in glob.glob(pat):
            fn = os.path.basename(f)
            if fn.startswith("~$"):
                continue
            if f in seen:
                continue
            seen.add(f)
            candidates.append(f)

    if not candidates:
        log("No Excel files found!", "ERROR")
        return None

    # Prefer files with JPDC/CRM name
    named = [f for f in candidates if "JPDC" in f.upper() or "CRM" in f.upper()]
    if named:
        candidates = named

    candidates.sort(key=os.path.getmtime, reverse=True)
    best = candidates[0]
    log(f"Selected: {best} (modified {datetime.fromtimestamp(os.path.getmtime(best)).strftime('%Y-%m-%d %H:%M')})")
    return best

def find_sheet(wb, keywords):
    """Find sheet matching any keyword in the list"""
    for kw in keywords:
        for name in wb.sheetnames:
            if kw.lower() in name.lower():
                return name
    return None

def find_header_row(rows, keywords, max_scan=10):
    """Scan first N rows to find header row containing keywords"""
    for i, row in enumerate(rows[:max_scan]):
        if row is None:
            continue
        text = " ".join([str(c) if c else "" for c in row[:12]])
        match_count = sum(1 for kw in keywords if kw in text)
        if match_count >= 2:
            return i
    return 0  # fallback to first row

def backup_db(conn):
    """Create timestamped backup"""
    backup_path = os.path.join(SOURCE_DIR, f"jpdc_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    shutil.copy2(DB_PATH, backup_path)
    log(f"Backup saved: {backup_path}")

    # Keep only last 10 backups
    backups = sorted(glob.glob(os.path.join(SOURCE_DIR, "jpdc_backup_*.db")))
    for old in backups[:-10]:
        os.remove(old)

# ── Table-specific import logic ─────────────────────────────
def import_units(ws, tower_val, cur, conn):
    """Import unit rows from Tower A or Tower E sheet"""
    rows = list(ws.iter_rows(values_only=True))
    log(f"  Sheet has {len(rows)} total rows")

    # Find header: the row where column B has numeric floor numbers
    # Tower sheets have format: row[0]=floor_num, row[1]=unit_no
    # Header detection by scanning for the first row where col[0] looks like a floor number
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        if row is None or len(row) < 2:
            continue
        r0 = str(row[0]).strip() if row[0] else ""
        r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        # Row 0 on Tower sheets usually has the title row like "TOWER A"
        # The data header has "楼层" / "单元号" / "Floor" / "Unit" etc
        if any(kw in r0 or kw in r1 for kw in ["Floor", "楼层", "楼栋", "Room", "单元号"]):
            header_idx = i
            break
        # If col[0] is purely numeric, this IS already a data row → header is row 0
        if r0.isdigit():
            header_idx = max(0, i - 1)  # the row above might be header
            break
    log(f"  Header at row {header_idx} (0-indexed)")

    # Sample header + first data row
    if header_idx < len(rows) and rows[header_idx]:
        h = [str(c)[:25] if c else "None" for c in rows[header_idx][:8]]
        log(f"  Header: {h}")
    if header_idx + 1 < len(rows) and rows[header_idx + 1]:
        d = [str(c)[:25] if c else "None" for c in rows[header_idx + 1][:8]]
        log(f"  First row: {d}")

    # Clear old data
    cur.execute(f"DELETE FROM units WHERE tower IN ('{tower_val}','TOWER {tower_val}')")
    log(f"  Cleared tower {tower_val}: {cur.rowcount} rows")

    # Import data rows
    count = 0
    skip_count = 0
    for i, row in enumerate(rows):
        if i <= header_idx:
            continue
        if not row or len(row) < 3:
            skip_count += 1
            continue

        # Data validation: col[0] must be a floor number (numeric)
        floor_raw = str(row[0]).strip() if row[0] else ""
        if not floor_raw or not floor_raw[0].isdigit():
            skip_count += 1
            continue

        vals = [clean(row[j]) if j < len(row) else None for j in range(14)]
        try:
            cur.execute("""INSERT INTO units (tower,floor_num,unit_no,unit_type,area_sqm,old_area_sqm,
                price_per_sqm,total_price,discount,discounted_price,status,situation,buyer_remarks,sales_date,remarks)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (tower_val,
                 str(vals[0]),           # floor_num
                 str(vals[1]) if vals[1] else "",  # unit_no
                 str(vals[2]) if vals[2] else None,  # unit_type
                 vals[3],                # area_sqm
                 vals[4],                # old_area_sqm
                 vals[5],                # price_per_sqm
                 vals[6],                # total_price
                 vals[7],                # discount
                 vals[8],                # discounted_price
                 str(vals[9]) if vals[9] else None,   # status
                 str(vals[10])[:500] if len(vals) > 10 and vals[10] else None,  # situation
                 str(vals[11])[:500] if len(vals) > 11 and vals[11] else None,  # buyer_remarks
                 str(vals[12]) if len(vals) > 12 and vals[12] else None,        # sales_date
                 str(vals[13])[:500] if len(vals) > 13 and vals[13] else None)) # remarks
            count += 1
        except Exception as e:
            log(f"  WARN: row {i} insert failed: {e}", "WARN")
            skip_count += 1

    if skip_count:
        log(f"  Skipped {skip_count} rows (empty/bad format)")
    log(f"  Imported {count} rows for Tower {tower_val}")
    conn.commit()
    return count

def import_overdue(ws, cur, conn):
    rows = list(ws.iter_rows(values_only=True))
    log(f"  Sheet has {len(rows)} total rows")

    header_idx = find_header_row(rows, ["买家", "单元", "逾期天数", "风险"])
    log(f"  Header at row {header_idx}")

    cur.execute("DELETE FROM overdue_warnings")
    count = 0
    for i, row in enumerate(rows):
        if i <= header_idx:
            continue
        if not row or not row[1]:
            continue
        try:
            cur.execute("""INSERT INTO overdue_warnings (buyer_name,tower,unit,paid_installments,total_paid,days_stopped,risk_level)
                VALUES (?,?,?,?,?,?,?)""",
                (str(row[1]), str(row[2]), str(row[3]),
                 int(row[4] or 0),
                 float(str(row[5]).replace(',', '')) if row[5] else 0,
                 int(row[6] or 0),
                 str(row[7]) if len(row) > 7 and row[7] else ''))
            count += 1
        except Exception as e:
            log(f"  WARN: overdue row {i} failed: {e}", "WARN")

    log(f"  Imported {count} overdue records")
    conn.commit()
    return count

def import_sold(ws, cur, conn):
    rows = list(ws.iter_rows(values_only=True))
    log(f"  Sheet has {len(rows)} total rows")

    header_idx = find_header_row(rows, ["楼栋", "单元号", "买家", "户型"])
    log(f"  Header at row {header_idx}")

    cur.execute("DELETE FROM sold_clients")
    count = 0
    for i, row in enumerate(rows):
        if i <= header_idx:
            continue
        if not row or len(row) < 7:
            continue
        unit_no = clean(row[2]) if len(row) > 2 else None
        if not unit_no or str(unit_no) in ("单元号(可点击)", "单元号", ""):
            continue

        vals = [clean(row[j]) if j < len(row) else None for j in range(12)]
        buyer = str(vals[6])[:200] if vals[6] else ''
        # skip header-like rows
        if buyer in ("买家/备注 Customer", ""):
            continue
        # skip tower header
        if str(vals[0]) in ("楼栋", ""):
            continue

        try:
            cur.execute("""INSERT INTO sold_clients (tower,floor_num,unit_no,unit_type,area_sqm,total_price,buyer_name,sd,sm,pia)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (str(vals[0]), str(vals[1]), str(vals[2]), str(vals[3]) if vals[3] else None,
                 vals[4], vals[5], buyer,
                 str(vals[7]) if vals[7] else None,
                 str(vals[8]) if vals[8] else None,
                 str(vals[9]) if vals[9] else None))
            count += 1
        except Exception as e:
            log(f"  WARN: sold row {i} failed: {e}", "WARN")

    log(f"  Imported {count} sold clients")
    conn.commit()
    return count

def import_payments(ws, cur, conn):
    rows = list(ws.iter_rows(values_only=True))
    log(f"  Sheet has {len(rows)} total rows")

    header_idx = find_header_row(rows, ["买家", "单元", "销售日期", "分期"])
    log(f"  Header at row {header_idx}")

    cur.execute("DELETE FROM payment_details")
    count = 0
    for i, row in enumerate(rows):
        if i <= header_idx:
            continue
        if not row or not row[0]:
            continue
        if str(row[0]).startswith("合计") or str(row[0]).startswith("Total"):
            continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(32)]
        try:
            cur.execute("""INSERT INTO payment_details (buyer_name,tower,unit_no,unit_type,sales_date,
                tr1_date,tr1_amount,tr2_date,tr2_amount,tr3_date,tr3_amount,tr4_date,tr4_amount,
                tr5_date,tr5_amount,tr6_date,tr6_amount,tr7_date,tr7_amount,tr8_date,tr8_amount)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (str(vals[0])[:100], str(vals[1])[:5], str(vals[3]) if vals[3] else '', '',
                 str(vals[4]) if vals[4] else '',
                 str(vals[5]) if vals[5] else '', vals[6],
                 str(vals[7]) if vals[7] else '', vals[8],
                 str(vals[9]) if vals[9] else '', vals[10],
                 str(vals[11]) if vals[11] else '', vals[12],
                 str(vals[13]) if vals[13] else '', vals[14],
                 str(vals[15]) if vals[15] else '', vals[16],
                 str(vals[17]) if vals[17] else '', vals[18],
                 str(vals[19]) if vals[19] else '', vals[20]))
            count += 1
        except Exception as e:
            log(f"  WARN: payment row {i} failed: {e}", "WARN")

    log(f"  Imported {count} payments")
    conn.commit()
    return count

def import_returned(ws, cur, conn):
    rows = list(ws.iter_rows(values_only=True))
    log(f"  Sheet has {len(rows)} total rows")

    header_idx = find_header_row(rows, ["楼栋", "单元号", "买家", "面积"])
    log(f"  Header at row {header_idx}")

    cur.execute("DELETE FROM returned_units")
    count = 0
    for i, row in enumerate(rows):
        if i <= header_idx:
            continue
        if not row or len(row) < 3 or not row[2]:
            continue
        unit_no = str(row[2]) if row[2] else ""
        if unit_no in ("单元号(可点击)", "单元号", ""):
            continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(11)]
        tower_val = str(vals[0]) if vals[0] else ""
        if tower_val in ("楼栋", ""):
            continue

        try:
            cur.execute("""INSERT INTO returned_units (tower,floor_num,unit_no,unit_type,area_sqm,price,buyer_notes,sd,sm,pia)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (tower_val, str(vals[1]), unit_no, str(vals[3]) if vals[3] else None,
                 vals[4], vals[5], str(vals[6])[:200] if vals[6] else '',
                 str(vals[7]) if vals[7] else None,
                 str(vals[8]) if vals[8] else None,
                 str(vals[9]) if vals[9] else None))
            count += 1
        except Exception as e:
            log(f"  WARN: returned row {i} failed: {e}", "WARN")

    log(f"  Imported {count} returned")
    conn.commit()
    return count

def import_problems(ws, cur, conn):
    rows = list(ws.iter_rows(values_only=True))
    log(f"  Sheet has {len(rows)} total rows")

    header_idx = find_header_row(rows, ["楼栋", "单元号", "问题", "面积"])
    log(f"  Header at row {header_idx}")

    cur.execute("DELETE FROM problem_units")
    count = 0
    for i, row in enumerate(rows):
        if i <= header_idx:
            continue
        if not row or len(row) < 3 or not row[2]:
            continue
        unit_no = str(row[2]) if row[2] else ""
        if unit_no in ("单元号(可点击)", "单元号", ""):
            continue
        vals = [clean(row[j]) if j < len(row) else None for j in range(12)]
        tower_val = str(vals[0]) if vals[0] else ""
        if tower_val in ("楼栋", ""):
            continue

        try:
            cur.execute("""INSERT INTO problem_units (tower,floor_num,unit_no,unit_type,area_sqm,price,buyer_notes,sd,sm,pia,issue_status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (tower_val, str(vals[1]), unit_no, str(vals[3]) if vals[3] else None,
                 vals[4], vals[5], str(vals[6])[:200] if vals[6] else '',
                 str(vals[7]) if vals[7] else None,
                 str(vals[8]) if vals[8] else None,
                 str(vals[9]) if vals[9] else None,
                 str(vals[10]) if len(vals) > 10 and vals[10] else None))
            count += 1
        except Exception as e:
            log(f"  WARN: problem row {i} failed: {e}", "WARN")

    log(f"  Imported {count} problems")
    conn.commit()
    return count

def rebuild_inventory(cur, conn):
    """Rebuild unit_inventory from units table"""
    cur.execute("DELETE FROM unit_inventory")
    for tower in ['A', 'E']:
        for utype in ['STUDIO', '2 BEDROOM', '3 BEDROOM', '4 BEDROOM', '5 BEDROOM', 'PENTHOUSE']:
            cur.execute("SELECT COUNT(*) FROM units WHERE tower=? AND unit_type=?", (tower, utype))
            total = cur.fetchone()[0]
            if total == 0:
                continue
            cur.execute("SELECT COUNT(*) FROM units WHERE tower=? AND unit_type=? AND status IN ('可售','可售 (NEW)')", (tower, utype))
            avail = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM units WHERE tower=? AND unit_type=? AND status IN ('已售','Sold','已售(LATE)')", (tower, utype))
            sold = cur.fetchone()[0]
            cur.execute("INSERT INTO unit_inventory (tower,unit_type,available,sold,total,updated_at) VALUES (?,?,?,?,?,?)",
                (tower, utype, avail, sold, total, datetime.now().isoformat()))
    conn.commit()
    log("  Inventory rebuilt")

# ── Main ────────────────────────────────────────────────────
def main():
    log("="*60)
    log("JPDC CRM - Data Import Started")
    log("="*60)

    excel_path = sys.argv[1] if len(sys.argv) > 1 else find_excel()
    if not excel_path:
        log("Aborted: no Excel file found", "ERROR")
        sys.exit(1)

    # Archive source Excel
    archive_name = f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(excel_path)}"
    archive_path = os.path.join(SOURCE_DIR, archive_name)
    shutil.copy2(excel_path, archive_path)
    log(f"Archived: {archive_name}")

    log(f"Opening: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    log(f"Sheets found: {wb.sheetnames}")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Backup
    backup_db(conn)

    results = {}
    all_ok = True

    # 1. Tower A
    sheet = find_sheet(wb, ["TOWER A", "Tower A"])
    if sheet:
        log(f"\n-- Tower A (sheet: {sheet}) --")
        results['Tower A'] = import_units(wb[sheet], 'A', cur, conn)
    else:
        log("WARN: Tower A sheet not found!", "WARN")

    # 2. Tower E
    sheet = find_sheet(wb, ["TOWER E", "Tower E"])
    if sheet:
        log(f"\n-- Tower E (sheet: {sheet}) --")
        results['Tower E'] = import_units(wb[sheet], 'E', cur, conn)
    else:
        log("WARN: Tower E sheet not found!", "WARN")

    # 3. Overdue
    sheet = find_sheet(wb, ["Overdue", "逾期"])
    if sheet:
        log(f"\n-- Overdue (sheet: {sheet}) --")
        results['Overdue'] = import_overdue(wb[sheet], cur, conn)
    else:
        log("WARN: Overdue sheet not found", "WARN")

    # 4. Sold Clients
    sheet = find_sheet(wb, ["Sold", "已售"])
    if sheet:
        log(f"\n-- Sold Clients (sheet: {sheet}) --")
        results['Sold'] = import_sold(wb[sheet], cur, conn)
    else:
        log("WARN: Sold sheet not found", "WARN")

    # 5. Payments
    sheet = find_sheet(wb, ["Payment", "付款"])
    if sheet:
        log(f"\n-- Payments (sheet: {sheet}) --")
        results['Payments'] = import_payments(wb[sheet], cur, conn)
    else:
        log("WARN: Payment sheet not found", "WARN")

    # 6. Returned
    sheet = find_sheet(wb, ["Returned", "退房"])
    if sheet:
        log(f"\n-- Returned (sheet: {sheet}) --")
        results['Returned'] = import_returned(wb[sheet], cur, conn)
    else:
        log("WARN: Returned sheet not found", "WARN")

    # 7. Problems
    sheet = find_sheet(wb, ["Problem", "问题"])
    if sheet:
        log(f"\n-- Problems (sheet: {sheet}) --")
        results['Problem'] = import_problems(wb[sheet], cur, conn)
    else:
        log("WARN: Problem sheet not found", "WARN")

    # Rebuild inventory
    log("\n-- Inventory --")
    rebuild_inventory(cur, conn)

    # Sync meta
    cur.execute("INSERT OR REPLACE INTO sync_meta (key, value, updated_at) VALUES ('last_import', ?, ?)",
        (datetime.now().isoformat(), datetime.now().isoformat()))
    cur.execute("INSERT OR REPLACE INTO sync_meta (key, value, updated_at) VALUES ('source_file', ?, ?)",
        (os.path.basename(excel_path), datetime.now().isoformat()))
    conn.commit()

    conn.close()
    wb.close()

    # Summary
    total = sum(results.values())
    log("\n" + "="*60)
    log("IMPORT SUMMARY")
    log("="*60)
    for k, v in results.items():
        log(f"  {k:<15}: {v:>6} records")
    log(f"  {'─'*25}")
    log(f"  {'TOTAL':<15}: {total:>6} records")
    log(f"  Source: {os.path.basename(excel_path)}")
    log(f"  Status: SUCCESS")
    log("="*60)

    log(f"\nImport completed. {total} records. Phone app: pull to refresh.")
    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        log(f"FATAL: {e}", "FATAL")
        import traceback
        log(traceback.format_exc(), "FATAL")
        sys.exit(1)
